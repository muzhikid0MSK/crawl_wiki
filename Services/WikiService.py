import openpyxl



import os
from PIL import Image
from concurrent.futures import ThreadPoolExecutor

import Services.Service
from Utils.CharUtils import validatetitle
import multiprocessing


from Utils import ThreadUtils
import logging, coloredlogs


class WikiService(Services.Service.Service):

    def __init__(self, excel_workbook: openpyxl.Workbook, browsers, browsers_locks, *args, **kwargs):
        super().__init__(*args, **kwargs)

        logging.getLogger(__name__)
        logging.basicConfig(format='%(asctime)s %(message)s')
        coloredlogs.install()

        self.res = {}
        self.excel_worksheet = excel_workbook.active
        self.browsers = browsers
        self.browsers_locks = browsers_locks
        self.url_prefix = ("https://zh.wikipedia.org/wiki/", "https://en.wikipedia.org/wiki/")
        self.ROOT = os.getcwd()
        self.thread_num = 1
        self._browser_index = None
        self._crawl_index = None
        if len(args) > 0:
            self.xlsx_name = args[0]
        else:
            self.xlsx_name = "unknown"
        self.person_name = kwargs.get("person", "")
        self.__to_crawl = []  # 先中后英
        self.__to_crawl_flag = []
        self.__pre()

    def __mkdir(self, path):
        if not os.path.exists(path):
            os.mkdir(path)

    def __pre(self):

        for r in range(1, self.excel_worksheet.max_row + 1):
            en_cell = self.excel_worksheet.cell(row=r, column=5)
            ch_cell = self.excel_worksheet.cell(row=r, column=6)
            state_cell = self.excel_worksheet.cell(row=r, column=4)
            self.__to_crawl.append((ch_cell.value, en_cell.value, state_cell.value,))

        military_type = self.xlsx_name[-2:]
        country_name = self.xlsx_name[:-2]
        self.__mkdir(os.path.join(self.ROOT, f"{self.person_name}"))
        self.__mkdir(os.path.join(self.ROOT, f"{self.person_name}", military_type))
        self.__mkdir(os.path.join(self.ROOT, f"{self.person_name}", military_type, country_name))
        self.result_root = os.path.join(self.ROOT, f"{self.person_name}", military_type, country_name)
        logging.info("初始化完成")

    def __save_as_html(self, index, html_path):
        res = self.browsers[index].execute_cdp_cmd('Page.captureSnapshot', {})

        # Write the file locally
        with open(html_path, 'w', newline='') as f:
            f.write(res['data'])

    def __handle_a_query(self, index_crawl):
        index = self._browser_index.increase_and_get()
        with self.browsers_locks[index]:
            for ch in range(2):
                search_flag = False
                logging.info(f"browser {index} is running")
                fullscreenshot_path = ""
                full_screen_img = None
                # to_crawl = validatetitle(self.__to_crawl[index_crawl][ch])
                state_name = validatetitle(self.__to_crawl[index_crawl][2])
                name_for_save = validatetitle(self.__to_crawl[index_crawl][0])  # 都存在中文目录下
                # 网址
                url = self.url_prefix[ch] + self.__to_crawl[index_crawl][ch]
                self.__mkdir(os.path.join(self.result_root, state_name))
                self.__mkdir(os.path.join(self.result_root, state_name, name_for_save))
                img_open = None
                try:
                    logging.info(f"---------------- 正在尝试在wiki上搜索{self.__to_crawl[index_crawl][ch]}  ----------------")
                    self.browsers[index].get(url)

                    element = self.browsers[index].find_element("xpath", "//*[@*='infobox vcard']")

                except:
                    logging.error(
                        f"---------------- There is no {self.__to_crawl[index_crawl][ch]} on wiki ! ----------------")
                else:
                    logging.info(f"----------- 已搜索到{self.__to_crawl[index_crawl][ch]} -----------")
                    search_flag = True
                    # self.__mkdir(os.path.join(self.result_root, name_for_save))
                    html_path = os.path.join(self.result_root, state_name, name_for_save,
                                             "information.mhtml")
                    logging.info(f"保存{self.__to_crawl[index_crawl][ch]} mhtml成功")
                    self.__save_as_html(index, html_path)
                    width = self.browsers[index].execute_script("return document.documentElement.scrollWidth")
                    height = self.browsers[index].execute_script("return document.body.scrollHeight")
                    # 将浏览器的宽高设置成刚刚获取的宽高
                    self.browsers[index].set_window_size(width, height)
                    # full screenshot
                    self.browsers[index].execute_script("window.print();")
                    # fullscreenshot_path = os.path.join(self.result_root,  to_crawl,
                    #                                    f"{validatetitle(self.browsers[i].title)}.jpg")
                    fullscreenshot_path = os.path.join(self.result_root, state_name, name_for_save,
                                                       "abstract.png")  # 实际上这里是暂存了整个页面的图片，取名abstract是为了之后真正存摘要图片是覆盖它
                    save_done = self.browsers[index].save_screenshot(fullscreenshot_path)

                    full_screen_pdf_path = os.path.join(self.result_root, state_name, name_for_save,
                                                        "information.pdf")
                    while not save_done:
                        pass
                    full_screen_img = Image.open(fullscreenshot_path)
                    if full_screen_img.mode in ("RGBA", "P"):
                        full_screen_img = full_screen_img.convert("RGB")
                    full_screen_img.save(full_screen_pdf_path)
                    logging.info(f"保存{self.__to_crawl[index_crawl][ch]} pdf 成功")
                    # html_path = os.path.join(self.result_root,  to_crawl,
                    #                          f"{validatetitle(self.browsers[i].title)}.html")
                    # with open(html_path, "w", encoding='utf-8') as f:
                    #     f.write(self.browsers[i].page_source)
                    # element shot
                    location = element.location
                    # to get the dimension the element
                    size = element.size
                    # to get the screenshot of complete page
                    # to get the x axis
                    x = location['x']
                    # to get the y axis
                    y = location['y']
                    # to get the length the element
                    height = location['y'] + size['height']
                    # to get the width the element
                    width = location['x'] + size['width']
                    # to open the captured image
                    img_open = Image.open(fullscreenshot_path)
                    # to crop the captured image to size of that element
                    img_open = img_open.crop((int(x), int(y), int(width), int(height)))
                    # to save the cropped image
                    element_shot_path = os.path.join(self.result_root, state_name, name_for_save,
                                                     "abstract.png")
                    if img_open.mode in ("RGBA", "P"):
                        img_open = img_open.convert("RGB")
                    img_open.save(element_shot_path)
                    logging.info(f"保存{self.__to_crawl[index_crawl][ch]} abstract 成功")
                finally:
                    if img_open is not None:
                        img_open.close()
                    if full_screen_img is not None:
                        full_screen_img.close()
                    if search_flag:
                        break

    def __del__(self):
        # logging.info(f"{self.xlsx_name} is closing ")
        # for browser in self.browsers:
        #     browser.close()
        logging.info(f"{self.xlsx_name} is closed ")

    def run(self, **params):
        super().run(**params)
        multi_thread = params.get("multi-thread", True)

        if multi_thread:
            self.thread_num = params.get("thread-num", multiprocessing.cpu_count() - 1)
        self._browser_index = ThreadUtils.ThreadSafeLoopCounter(self.thread_num)

        # url_prefix = "https://baike.baidu.com/item/"
        self.res["urls"] = []



        with ThreadPoolExecutor(self.thread_num) as pool:

            for i in range(len(self.__to_crawl)):
                pool.submit(self.__handle_a_query, i)
